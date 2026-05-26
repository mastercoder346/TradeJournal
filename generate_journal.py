import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule

def create_trading_journal():
    wb = openpyxl.Workbook()
    
    # -------------------------------------------------------------
    # SHEET 1: DASHBOARD
    # -------------------------------------------------------------
    ws_dash = wb.active
    ws_dash.title = "Dashboard"
    ws_dash.views.sheetView[0].showGridLines = True
    
    # Premium Dark Colors
    fill_dark_bg = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid") # Slate-900
    fill_card = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid") # Slate-800
    fill_header = PatternFill(start_color="334155", end_color="334155", fill_type="solid") # Slate-700
    
    font_title = Font(name="Segoe UI", size=18, bold=True, color="FFFFFF")
    font_subtitle = Font(name="Segoe UI", size=10, italic=True, color="94A3B8")
    font_label = Font(name="Segoe UI", size=9, bold=True, color="94A3B8")
    font_value = Font(name="Segoe UI", size=20, bold=True, color="F8FAFC")
    font_value_green = Font(name="Segoe UI", size=20, bold=True, color="10B981")
    
    # Grid lines/Borders
    thin_border = Border(
        left=Side(style='thin', color='475569'),
        right=Side(style='thin', color='475569'),
        top=Side(style='thin', color='475569'),
        bottom=Side(style='thin', color='475569')
    )
    
    # Dashboard Header
    ws_dash["B2"] = "TRADING PERFORMANCE DASHBOARD"
    ws_dash["B2"].font = font_title
    ws_dash["B3"] = "Real-time updates dynamically computed from your Trade Log sheet"
    ws_dash["B3"].font = font_subtitle
    
    # KPI 1: Net P&L Card
    ws_dash.merge_cells("B5:C5")
    ws_dash["B5"] = "TOTAL NET PROFIT"
    ws_dash["B5"].font = font_label
    ws_dash["B5"].fill = fill_card
    ws_dash["B5"].alignment = Alignment(horizontal="center")
    
    ws_dash.merge_cells("B6:C6")
    ws_dash["B6"] = "='Trade Log'!M3 + SUM('Trade Log'!M:M)"  # Just placeholder for formula
    ws_dash["B6"] = "=SUM('Trade Log'!M:M)"
    ws_dash["B6"].font = font_value_green
    ws_dash["B6"].fill = fill_card
    ws_dash["B6"].alignment = Alignment(horizontal="center")
    ws_dash["B6"].number_format = "$#,##0.00;($#,##0.00);\"-\""
    
    # KPI 2: Win Rate
    ws_dash.merge_cells("E5:F5")
    ws_dash["E5"] = "WIN RATE"
    ws_dash["E5"].font = font_label
    ws_dash["E5"].fill = fill_card
    ws_dash["E5"].alignment = Alignment(horizontal="center")
    
    ws_dash.merge_cells("E6:F6")
    ws_dash["E6"] = "=IF(COUNTA('Trade Log'!C:C)-1>0, COUNTIF('Trade Log'!M:M,\">0\") / (COUNTIF('Trade Log'!M:M,\">0\") + COUNTIF('Trade Log'!M:M,\"<0\")), 0)"
    ws_dash["E6"].font = font_value
    ws_dash["E6"].fill = fill_card
    ws_dash["E6"].alignment = Alignment(horizontal="center")
    ws_dash["E6"].number_format = "0.0%"
    
    # KPI 3: Total Trades
    ws_dash.merge_cells("H5:I5")
    ws_dash["H5"] = "TOTAL TRADES"
    ws_dash["H5"].font = font_label
    ws_dash["H5"].fill = fill_card
    ws_dash["H5"].alignment = Alignment(horizontal="center")
    
    ws_dash.merge_cells("H6:I6")
    ws_dash["H6"] = "=COUNTA('Trade Log'!C:C)-1"
    ws_dash["H6"].font = font_value
    ws_dash["H6"].fill = fill_card
    ws_dash["H6"].alignment = Alignment(horizontal="center")
    ws_dash["H6"].number_format = "#,##0"

    # KPI 4: Profit Factor
    ws_dash.merge_cells("K5:L5")
    ws_dash["K5"] = "PROFIT FACTOR"
    ws_dash["K5"].font = font_label
    ws_dash["K5"].fill = fill_card
    ws_dash["K5"].alignment = Alignment(horizontal="center")
    
    ws_dash.merge_cells("K6:L6")
    ws_dash["K6"] = "=IF(SUMIF('Trade Log'!M:M,\"<0\")<>0, ABS(SUMIF('Trade Log'!M:M,\">0\") / SUMIF('Trade Log'!M:M,\"<0\")), SUMIF('Trade Log'!M:M,\">0\"))"
    ws_dash["K6"].font = font_value
    ws_dash["K6"].fill = fill_card
    ws_dash["K6"].alignment = Alignment(horizontal="center")
    ws_dash["K6"].number_format = "0.00"

    # Apply borders and styling to the KPI blocks
    for col in ["B", "C", "E", "F", "H", "I", "K", "L"]:
        ws_dash[f"{col}5"].border = thin_border
        ws_dash[f"{col}6"].border = thin_border

    # -------------------------------------------------------------
    # SHEET 2: TRADE LOG
    # -------------------------------------------------------------
    ws_log = wb.create_sheet(title="Trade Log")
    ws_log.views.sheetView[0].showGridLines = True
    
    # Premium Light-Medium styling for log
    fill_log_hdr = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid") # Dark headers
    font_log_hdr = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
    
    headers = [
      "Date", "Asset Class", "Ticker", "Direction", "Option Strategy", 
      "Strike Price(s)", "Expiry Date", "Quantity", "Entry Price", "Exit Price", 
      "Fees & Comm.", "Gross P&L", "Emotion Pre", "Emotion Post", "Setup Reason", "Notes"
    ]
    
    for col_num, header in enumerate(headers, 1):
        cell = ws_log.cell(row=1, column=col_num)
        cell.value = header
        cell.font = font_log_hdr
        cell.fill = fill_log_hdr
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws_log.row_dimensions[1].height = 28
    
    # Pre-populate with Seed Trades (showing Spreads, Calendars, etc.)
    seed_rows = [
        ["2026-05-10 10:30", "Stock", "AAPL", "Buy/Long", "N/A", "", "", 100, 172.50, 178.20, 4.95, "= (J2 - I2) * H2 - K2", "Calm", "Calm", "Breakout", "Perfect textbook breakout of the 50 DMA."],
        ["2026-05-12 14:15", "Option", "TSLA", "Buy/Long", "Single Option", "220 Call", "2026-06-15", 5, 4.50, 6.20, 7.50, "= (J3 - I3) * H3 * 100 - K3", "FOMO", "Greed", "Trend Following", "Surged up. Chased slightly due to FOMO."],
        ["2026-05-14 09:50", "Option", "SPY", "Sell/Short", "Credit Spread", "510/505 Put Spread", "2026-05-29", 10, 1.20, 0.30, 12.00, "= (I4 - J4) * H4 * 100 - K4", "Calm", "Calm", "Support/Resistance", "Sold put credit spread at solid horizontal support. Worked perfectly."],
        ["2026-05-18 11:30", "Option", "MSFT", "Buy/Long", "Debit Spread", "420/425 Call Spread", "2026-06-19", 4, 2.10, 3.40, 8.00, "= (J5 - I5) * H5 * 100 - K5", "Calm", "Stressed", "Trend Following", "Long term call debit spread capturing upward continuation."],
        ["2026-05-20 13:20", "Option", "NVDA", "Buy/Long", "Calendar Spread", "920 Call (May/Jun)", "2026-06-19", 2, 4.80, 5.50, 5.00, "= (J6 - I6) * H6 * 100 - K6", "Stressed", "Calm", "News/Catalyst", "Traded pre-earnings calendar spread. Implied volatility crush arbitrage."],
        ["2026-05-22 15:40", "Stock", "AMD", "Buy/Long", "N/A", "", "", 50, 165.00, 158.50, 0.00, "= (J7 - I7) * H7 - K7", "Fear", "Stressed", "FOMO (No Setup)", "Traded without clear setup feeling extremely anxious. Stop hit."]
    ]
    
    for row_idx, row_data in enumerate(seed_rows, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws_log.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.font = Font(name="Segoe UI", size=10)
            
            # Format numbers
            if col_idx in [8, 9, 10, 11]: # Qty, Prices, Fees
                if value != "":
                    cell.value = float(value)
            
            # Formatting styles
            if col_idx in [9, 10, 11]:
                cell.number_format = "$#,##0.00"
            elif col_idx == 8:
                cell.number_format = "#,##0"
                
        # Setup Net P&L cell formatting
        pnl_cell = ws_log.cell(row=row_idx, column=12)
        pnl_cell.number_format = "$#,##0.00;($#,##0.00);\"-\""

    # Configure Dropdown Valdiations
    dv_asset = DataValidation(type="list", formula1='"Stock,Option"', allow_blank=True)
    dv_action = DataValidation(type="list", formula1='"Buy/Long,Sell/Short"', allow_blank=True)
    dv_strategy = DataValidation(type="list", formula1='"N/A,Single Option,Credit Spread,Debit Spread,Calendar Spread,Diagonal Spread,Iron Condor,Covered Call,Cash Secured Put"', allow_blank=True)
    dv_emotion = DataValidation(type="list", formula1='"Calm,FOMO,Greed,Fear,Stressed,Neutral"', allow_blank=True)
    
    ws_log.add_data_validation(dv_asset)
    ws_log.add_data_validation(dv_action)
    ws_log.add_data_validation(dv_strategy)
    ws_log.add_data_validation(dv_emotion)
    
    # Apply validations to 100 rows
    dv_asset.add("B2:B100")
    dv_action.add("D2:D100")
    dv_strategy.add("E2:E100")
    dv_emotion.add("M2:N100")

    # Add Conditional formatting to Net P&L (Emerald Green vs Coral Red)
    green_fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid") # Soft emerald
    red_fill = PatternFill(start_color="FFE4E6", end_color="FFE4E6", fill_type="solid") # Soft coral
    
    ws_log.conditional_formatting.add("L2:L100", CellIsRule(operator="greaterThan", formula=["0"], fill=green_fill, font=Font(color="065F46", bold=True)))
    ws_log.conditional_formatting.add("L2:L100", CellIsRule(operator="lessThan", formula=["0"], fill=red_fill, font=Font(color="991B1B", bold=True)))

    # -------------------------------------------------------------
    # SHEET 3: PSYCHOLOGY AUDIT
    # -------------------------------------------------------------
    ws_audit = wb.create_sheet(title="Psychology & Setup Audit")
    ws_audit.views.sheetView[0].showGridLines = True
    
    ws_audit["B2"] = "PSYCHOLOGICAL PERFORMANCE BREAKDOWN"
    ws_audit["B2"].font = Font(name="Segoe UI", size=14, bold=True, color="1E293B")
    
    # Emotion Audit Headers
    ws_audit["B4"] = "Pre-Trade Emotion"
    ws_audit["C4"] = "Trades Count"
    ws_audit["D4"] = "Win Rate"
    ws_audit["E4"] = "Total P&L"
    
    for col in ["B", "C", "D", "E"]:
        ws_audit[f"{col}4"].font = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
        ws_audit[f"{col}4"].fill = fill_log_hdr
        ws_audit[f"{col}4"].alignment = Alignment(horizontal="center")
    
    emotions = ["Calm", "FOMO", "Greed", "Fear", "Stressed", "Neutral"]
    for idx, emo in enumerate(emotions, 5):
        ws_audit[f"B{idx}"] = emo
        ws_audit[f"B{idx}"].font = Font(name="Segoe UI", size=10)
        
        # Trades Count formula
        ws_audit[f"C{idx}"] = f"=COUNTIF('Trade Log'!M:M, \"{emo}\")"
        ws_audit[f"C{idx}"].font = Font(name="Segoe UI", size=10)
        ws_audit[f"C{idx}"].number_format = "#,##0"
        
        # Win Rate Formula
        ws_audit[f"D{idx}"] = f"=IF(C{idx}>0, COUNTIFS('Trade Log'!M:M, \"{emo}\", 'Trade Log'!L:L, \">0\") / C{idx}, 0)"
        ws_audit[f"D{idx}"].font = Font(name="Segoe UI", size=10)
        ws_audit[f"D{idx}"].number_format = "0.0%"
        
        # Total P&L Formula
        ws_audit[f"E{idx}"] = f"=SUMIF('Trade Log'!M:M, \"{emo}\", 'Trade Log'!L:L)"
        ws_audit[f"E{idx}"].font = Font(name="Segoe UI", size=10, bold=True)
        ws_audit[f"E{idx}"].number_format = "$#,##0.00;($#,##0.00);\"-\""

    # Adjust auto widths for all sheets
    for ws in [ws_dash, ws_log, ws_audit]:
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
            
    # Save spreadsheet
    file_path = "Trading_Journal_v1.xlsx"
    wb.save(file_path)
    print(f"Spreadsheet constructed successfully at {file_path}")

if __name__ == "__main__":
    create_trading_journal()
